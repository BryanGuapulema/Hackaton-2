# lambda_function.py
import os, io, csv, re, boto3
from datetime import datetime, timezone
from openpyxl import load_workbook

s3 = boto3.client('s3')
dynamodb = boto3.resource('dynamodb')

BUCKET = os.environ['BUCKET_NAME']
INBOX_PREFIX = os.environ['EXCEL_SOURCE_PREFIX']
CONTROL_TABLE = os.environ['CONTROL_TABLE']
ERROR_TABLE = os.environ['ERROR_TABLE']
SHEET_NAME = os.environ.get('EXCEL_SHEET_NAME', 'storesBudget')

def _now_iso(): return datetime.now(timezone.utc).isoformat()

def _clean_money(v):
    """
    Normaliza valores monetarios sin perder magnitud.
    Ejemplos:
      "$60.749.820.000"    -> "60749820000"
      "60.749.820,50"      -> "60749820.50"
      60749820000 (num)    -> "60749820000"
      "  60 749 820 000 "  -> "60749820000"
    """
    import re
    if v is None:
        return None
    # Si openpyxl ya entrega número, no fuerces a float (pierde formato), solo int si es entero
    if isinstance(v, (int, float)):
        # si es entero exacto, devuélvelo como entero
        if float(v).is_integer():
            return str(int(v))
        else:
            # número con decimales reales
            return f"{v}"

    s = str(v).strip()
    if not s:
        return None

    # Mantén solo dígitos, coma y punto (para decidir decimal)
    s = re.sub(r"[^\d,\.]", "", s)

    # Caso con ambos separadores: toma el ÚLTIMO como decimal, el resto son miles
    if ("," in s) and ("." in s):
        last = max(s.rfind(","), s.rfind("."))
        int_part = re.sub(r"\D", "", s[:last])           # quita miles
        dec_part = re.sub(r"\D", "", s[last+1:]) or "0"  # decimales (si hay)
        return f"{int_part}.{dec_part}"

    # Solo un tipo de separador
    if "," in s:
        # Asume coma decimal (estilo es-ES). Quita puntos/espacios y cambia coma por punto.
        s = s.replace(".", "")
        s = s.replace(",", ".")
        return s
    if "." in s:
        # Ambiguo: podría ser miles o decimal. Si hay exactamente 3 dígitos tras el punto repetidos (miles),
        # elimina todos los puntos (asume solo miles). Si parece decimal (p.ej. 123.45), deja el punto.
        parts = s.split(".")
        if all(len(p) == 3 for p in parts[1:]):  # patrón de miles 1.234.567
            return "".join(parts)
        else:
            return s  # probablemente decimal

    # Sin separadores: deja solo dígitos
    digits = re.sub(r"\D", "", s)
    return digits or None

    """Convierte valores tipo '$60.749.820.000' o '60.749.820,50' a '60749820000' o '60749820.50'."""
    if v is None: return None
    # si openpyxl ya trajo número
    if isinstance(v, (int, float)):
        return str(v)
    v = str(v).strip()
    if not v: return None
    # quita símbolos comunes
    v = re.sub(r'[\s$\€\£]', '', v)
    # si tiene una coma decimal (estilo es-ES), elimina puntos miles y cambia coma por punto
    if v.count(',') == 1 and (v.replace(',', '').replace('.', '').isdigit()):
        v = v.replace('.', '').replace(',', '.')
    else:
        # sin coma decimal, elimina puntos de miles
        v = v.replace('.', '')
    # valida final
    return v if re.fullmatch(r'-?\d+(\.\d+)?', v) else None

def lambda_handler(event, context):
    run_month = event.get("run_month")  # "YYYY-MM"
    if not run_month: raise RuntimeError("Falta run_month en el evento")

    run_id = f"budget_{run_month}_{int(datetime.utcnow().timestamp())}"

    # 1) localizar el primer .xlsx en el inbox
    listed = s3.list_objects_v2(Bucket=BUCKET, Prefix=INBOX_PREFIX)
    objects = listed.get('Contents', [])
    xlsx_keys = [o['Key'] for o in objects if o['Key'].lower().endswith('.xlsx')]
    if not xlsx_keys:
        raise RuntimeError(f"No se encontró .xlsx en s3://{BUCKET}/{INBOX_PREFIX}")
    src_key = sorted(xlsx_keys)[0]

    # 2) leer el .xlsx en memoria
    body = s3.get_object(Bucket=BUCKET, Key=src_key)['Body'].read()
    wb = load_workbook(io.BytesIO(body), data_only=True)
    ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.worksheets[0]

    # 3) detectar encabezado (primera fila no vacía) y mapear columnas
    # busca fila con al menos 2 celdas no vacías
    header_row_idx = None
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row and sum(1 for c in row if c not in (None, "")) >= 2:
            header_row_idx = i
            break
    if header_row_idx is None:
        raise RuntimeError("No se encontró fila de encabezado en la hoja")

    headers = [str(c).strip().lower() if c is not None else "" 
               for c in next(ws.iter_rows(min_row=header_row_idx, max_row=header_row_idx, values_only=True))]
    # funciones para extraer StoreID y Budget
    def _idx(options):
        for name in options:
            if name in headers:
                return headers.index(name)
        return None

    idx_sid = _idx(["storeid","store_id","id","tienda","store"])
    idx_bud = _idx(["budget","presupuesto","monto","importe"])
    if idx_sid is None or idx_bud is None:
        raise RuntimeError(f"No se hallaron columnas StoreID/Budget en encabezados: {headers}")

    # 4) recorrer filas de datos (después del encabezado) y normalizar
    out = io.StringIO()
    writer = csv.DictWriter(out, fieldnames=['StoreID','Budget'])
    writer.writeheader()
    count = 0

    for row in ws.iter_rows(min_row=header_row_idx+1, values_only=True):
        if row is None: continue
        sid = row[idx_sid] if idx_sid < len(row) else None
        bud = row[idx_bud] if idx_bud < len(row) else None
        if sid in (None, "") or bud in (None, ""): 
            continue
        # normaliza valores
        sid_str = str(sid).strip()
        bud_clean = _clean_money(bud)
        if not sid_str or not bud_clean: 
            continue

        if count < 5:
            print("DEBUG Budget raw/clean:", bud, "->", bud_clean)

        writer.writerow({'StoreID': sid_str, 'Budget': bud_clean})
        count += 1

    # 5) escribir salida en Bronze/run_month
    dest_prefix = f"bronze/source=excel/table=storesBudget/csv/"
    dest_key = f"{dest_prefix}storesBudget.csv"
    s3.put_object(Bucket=BUCKET, Key=dest_key, Body=out.getvalue().encode('utf-8'))

    # 6) control
    dynamodb.Table(CONTROL_TABLE).put_item(Item={
        "run_id": run_id, "run_month": run_month, "source":"excel", "table":"storesBudget",
        "status":"SUCCEEDED", "records_out": count,
        "started_at": _now_iso(), "ended_at": _now_iso()
    })

    return {"run_id": run_id, "records_out": count, "dest_key": dest_key}
